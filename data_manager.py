from openpyxl import load_workbook
from datetime import datetime
import os

# this function converts utc time format to usual one
def utc_to_local(utc_time:str):
    ts = int(utc_time)
    return datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')

# this class is created for managing flight info and writing it to Excel file
class DataManager:
    def __init__(self, data):
        self.clear_excel()
        self.data = data
        self.write_to_exel()

# This function clears Excel file
    def clear_excel(self):
        file = "Information.xlsx"
        wb = load_workbook(file)
        excel_list = wb["Flights Data"]
        excel_list.delete_cols(1, 7)
        excel_list.delete_rows(1, 10000)
        wb.save(file)
        wb.close()

    # This function writes information to Excel and opens Excel file
    def write_to_exel(self):
        file = "Information.xlsx"
        wb = load_workbook(file)
        excel_list = wb["Flights Data"]
        if self.data["_results"] == 0:
            excel_list.append(["There are no flights available"])
        else:
            excel_list.append(["Route", "",
                               "Fly Duration",
                               "Departure Time",
                               "Arrival time",
                               "Starting price (Euro)",
                               "Link for Booking,"])
            for item in self.data["data"]:
                excel_list.append([item["cityFrom"],
                                   item["cityTo"],
                                   item["fly_duration"],
                                   utc_to_local(item["dTime"]),
                                   utc_to_local(item["aTime"]),
                                   item['price'],
                                   item["deep_link"]])
                if len(item["route"]) > 1:
                    excel_list.append(["Route:"])
                    for part in item["route"]:
                        excel_list.append([part["cityFrom"],
                                           part["cityTo"],
                                           " ",
                                           utc_to_local(part["dTime"]),
                                           utc_to_local(part["aTime"])])
                excel_list.append(["------", "------", "------"])
        wb.save(file)
        wb.close()
        os.system("open Information.xlsx")
        print("Success!")
